import os
import time
import pyvisa as visa
import serial
import serial.tools.list_ports_windows
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from pyvisa import VisaIOError
from serial import SerialException


def powerSetVolt(power, Ptype, volt):
    # sysTem = power.query('*IDN?')
    if 'DH1766' in Ptype:
        # power.write('INST CH1\n')
        power.write('APPL:VOLT %f,%f\n' % (volt, volt))
    else:
        power.write('VOLT:RANG 6,(@1)\n')
        power.write('VOLT %f,(@1)\n' % volt)


def powerSetCurrent(power, Ptype, current):
    # sysTem = power.query('*IDN?')
    if 'DH1766' in Ptype:
        # power.write('INST CH1\n')
        power.write('APPL:CURR %f,%f\n' % (current, current))
    else:
        power.write('CURR:RANG 0.5,(@1)\n')
        power.write('CURR %f,(@1)\n' % current)


def powerON(power, Ptype):
    # sysTem = power.query('*IDN?')
    if 'DH1766' in Ptype:
        # power.write('INST CH1\n')
        power.write('APPL:OUTP ON,ON\n')
    else:
        power.write('OUTP ON,(@1)\n')


def powerOff(power, Ptype):
    # sysTem = power.query('*IDN?')
    if 'DH1766' in Ptype:
        # power.write('INST CH1\n')
        power.write('APPL:OUTP OFF,OFF\n')
    else:
        power.write('OUTP OFF,(@1)\n')


def getWorkBook() -> Workbook:
    if not os.path.exists('./result.xlsx'):
        wkbook = Workbook()
        wksheet = wkbook.active
        wksheet.merge_cells('B1:C1')
        wksheet.merge_cells('D1:E1')
        wksheet.merge_cells('F1:G1')
        wksheet.merge_cells('H1:I1')
        wksheet['A1'] = '电源电压'
        wksheet['B1'] = 'adc value'
        wksheet['D1'] = 'voltage'
        wksheet['F1'] = 'adc value'
        wksheet['H1'] = 'voltage'
        wksheet.append(['', 'Min', 'Max',  'Min', 'Max', 'Min', 'Max', 'Min', 'Max'])
        wkbook.save('result.xlsx')
    wkbook = load_workbook('result.xlsx')
    return wkbook


def readData2Excel(dataList, com):
    result_list = dataList
    print('测试串口' + com)
    adc_value_list = []
    vol_list = []
    try:
        sercom = serial.Serial(com, 115200)
        sercom.flushOutput()
        time.sleep(1)
        while True:
            if sercom.inWaiting():
                recv_data = sercom.readline().decode('utf-8', errors='replace')
                if 'ntc_drv_read_temp adc value' in recv_data:
                    adc_value = recv_data.split(':')[-1].strip()
                    adc_value_list.append(int(adc_value))
                    print(recv_data)
                elif 'ntc_drv_read_temp cali voltage value' in recv_data:
                    vol = recv_data.split(':')[-1].strip().split(' ')[0]
                    vol_list.append(int(vol))
                    print(recv_data)
                if len(adc_value_list) == 20 or len(vol_list) == 20:
                    result_list.append(min(adc_value_list))
                    result_list.append(max(adc_value_list))
                    result_list.append(min(vol_list))
                    result_list.append(max(vol_list))
                    break
        sercom.close()
    except SerialException:
        print('串口开启失败！')
        time.sleep(2)
    except UnicodeDecodeError:
        print('decode error')
    print(result_list)
    workbook = getWorkBook()
    workbook.active.append(result_list)
    workbook.save('result.xlsx')


class ADCTest:

    def __init__(self):
        self.work_path = os.getcwd()
        self.config_file = open(self.work_path + "\\config.txt", encoding='utf-8')
        content = self.config_file.read()
        self.address = content.split('\n')[0].split('=')[1]
        self.powerType = content.split('\n')[1].split('=')[1]
        self.start = content.split('\n')[2].split('=')[1]
        self.end = content.split('\n')[3].split('=')[1]
        self.step = content.split('\n')[4].split('=')[1]
        # self.number = content.split('\n')[5].split('=')[1]

    def startTest(self):
        try:
            rm = visa.ResourceManager()
            powerSuppler = rm.open_resource(self.address)
        except VisaIOError:
            print('无法连接电源')
            time.sleep(2)
        else:
            if os.path.exists('result.xlsx'):
                os.remove('result.xlsx')
            powerSetVolt(powerSuppler, self.powerType, 0)
            powerSetCurrent(powerSuppler, self.powerType, 0.1)
            powerON(powerSuppler, self.powerType)

            portlist = serial.tools.list_ports_windows.comports()
            com = str(portlist[0]).split('-')[0]
            print('开始测试')
            time.sleep(2)
            for testvol in range(int(self.start), int(self.end) + int(self.step), int(self.step)):
                voltage = testvol / 1000
                print('测试电压%f' % voltage)
                powerSetVolt(powerSuppler, self.powerType, voltage)
                time.sleep(1)
                dataList = [testvol]
                readData2Excel(dataList, com)
            powerOff(powerSuppler, self.powerType)


if __name__ == '__main__':
    ADCTest().startTest()
