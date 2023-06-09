import os
import threading
import time
import pyvisa as visa
import serial
import serial.tools.list_ports_windows
from openpyxl.workbook import Workbook
from pyvisa import VisaIOError
from serial import SerialException


"""该程序以多线程方式去读取串口数据，支持同时多个串口，提高测试效率"""


def powerSetVolt(power, Ptype, volt):
    if 'DH1766' in Ptype:
        power.write('APPL:VOLT %f,%f\n' % (volt, volt))
    else:
        power.write('VOLT:RANG 6,(@1)\n')
        power.write('VOLT %f,(@1)\n' % volt)


def powerSetCurrent(power, Ptype, current):
    if 'DH1766' in Ptype:
        power.write('APPL:CURR %f,%f\n' % (current, current))
    else:
        power.write('CURR:RANG 0.5,(@1)\n')
        power.write('CURR %f,(@1)\n' % current)


def powerON(power, Ptype):
    if 'DH1766' in Ptype:
        power.write('APPL:OUTP ON,ON\n')
    else:
        power.write('OUTP ON,(@1)\n')


def powerOff(power, Ptype):
    if 'DH1766' in Ptype:
        power.write('APPL:OUTP OFF,OFF\n')
    else:
        power.write('OUTP OFF,(@1)\n')


class runThread(threading.Thread):
    """
    该线程用于读取串口数据，并提取ADC结果保存到excel。创建线程需要串口和测试步长2个参数。
    """

    def __init__(self, comport, step):
        super().__init__()
        self.comport = comport
        self.step = step
        self.running = True
        self.completeFlag = False

    def stopThread(self):
        self.running = False

    def changeState(self):
        self.completeFlag = False

    def run(self) -> None:
        serialCom = serial.Serial(self.comport, 115200)
        workbook = getWorkBook(self.comport)
        adc_value_list = []
        vol_list = []
        result_list = []
        voltage = 0
        while self.running:
            if not self.completeFlag:
                if not serialCom.isOpen():
                    serialCom.open()
                if serialCom.inWaiting():
                    recv_data = serialCom.readline().decode('utf-8', errors='replace')
                    print(recv_data)
                    if 'user_app_main adc value' in recv_data:
                        adc_value = recv_data.split(':')[1].strip()
                        adc_value_list.append(int(adc_value))
                    elif 'user_app_main cali voltage value' in recv_data:
                        vol = recv_data.split(':')[1].strip().split(' ')[0]
                        vol_list.append(int(vol))
                    if len(adc_value_list) == 10 and len(vol_list) == 10:
                        result_list.append(voltage)
                        result_list.append(min(adc_value_list))
                        result_list.append(max(adc_value_list))
                        result_list.append(min(vol_list))
                        result_list.append(max(vol_list))
                        print(result_list)
                        workbook.active.append(result_list)
                        workbook.save(self.comport + 'result.xlsx')
                        result_list.clear()
                        adc_value_list.clear()
                        vol_list.clear()
                        voltage += int(self.step)
                        self.completeFlag = True
                        serialCom.close()
                else:
                    time.sleep(0.1)
            else:
                time.sleep(0.2)


def getWorkBook(filename) -> Workbook:
    wkbook = Workbook()
    wksheet = wkbook.active
    wksheet.merge_cells('B1:C1')
    wksheet.merge_cells('D1:E1')
    wksheet['A1'] = '电源电压'
    wksheet['B1'] = 'adc value'
    wksheet['D1'] = 'voltage'
    wksheet.append(['', 'Min', 'Max', 'Min', 'Max'])
    wkbook.save(filename + 'result.xlsx')
    return wkbook


class ADCTest:

    def __init__(self):
        self.work_path = os.getcwd()
        self.config_file = open(self.work_path + "\\config.txt", encoding='utf-8')
        content = self.config_file.read()
        self.address = content.split('\n')[0].split('=')[1]
        self.powerType = content.split('\n')[1].split('=')[1]
        self.start = content.split('\n')[2].split('=')[1]
        self.end = content.split('\n')[3].split('=')[1]
        self.step = content.split('\n')[4].split('=')[1]

    def startTest(self):
        try:
            rm = visa.ResourceManager()
            powerSuppler = rm.open_resource(self.address)
        except VisaIOError:
            print('无法连接电源')
            time.sleep(2)
        else:
            powerSetVolt(powerSuppler, self.powerType, 0)
            powerSetCurrent(powerSuppler, self.powerType, 0.1)
            powerON(powerSuppler, self.powerType)
            print('开始测试')
            print(time.strftime('%H:%M:%S', time.localtime()))

            threadList = []
            com_list = serial.tools.list_ports_windows.comports()
            for item in com_list:   # 根据串口数去创建excel文件和线程
                com = str(item).split('-')[0]
                if os.path.exists(com + 'result.xlsx'):
                    os.remove(com + 'result.xlsx')
                thread1 = runThread(com, self.step)
                threadList.append(thread1)
                thread1.start()
            for testvol in range(int(self.start), int(self.end) + int(self.step), int(self.step)):
                voltage = testvol / 1000
                print('测试电压%f' % voltage)
                powerSetVolt(powerSuppler, self.powerType, voltage)
                time.sleep(11)
                for my_thread in threadList:
                    my_thread.changeState()     # 将线程中已完成标准置为False，以便开始下一次测试
            powerOff(powerSuppler, self.powerType)
            print('测试结束')
            print(time.strftime('%H:%M:%S', time.localtime()))
            for my_thread in threadList:
                my_thread.stopThread()


if __name__ == '__main__':
    ADCTest().startTest()
